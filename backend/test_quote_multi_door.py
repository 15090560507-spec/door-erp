import io
import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BACKEND_DIR)

from models import CADRequest
from quote_database import AccessoryDatabaseManager, QuoteDatabaseManager
from quote_excel import generate_excel


def _multi_door_quote():
    return {
        "customerName": "多门客户",
        "projectName": "",
        "quoteDate": "2026-07-23",
        "noticeText": "本报价不含税工厂结算价，含木箱。",
        "doorGroups": [
            {
                "groupName": "入户门",
                "taskId": "task-1",
                "pricingMode": "outerArea",
                "trimUnitPrice": 0,
                "items": [
                    {
                        "category": "门类组合",
                        "productName": "子母门 0.8纯铜 紫荆花款 竖条款",
                        "width": 1600,
                        "height": 2600,
                        "openDirection": "右外开",
                        "unit": "m2",
                        "unitPrice": 1500,
                    },
                    {
                        "category": "合页",
                        "productName": "暗合页（6只）",
                        "quantity": 1,
                        "unit": "套",
                        "unitPrice": 1000,
                    },
                ],
            },
            {
                "groupName": "厨房门",
                "taskId": "task-2",
                "pricingMode": "framePlusTrim",
                "trimUnitPrice": 200,
                "items": [
                    {
                        "category": "门类组合",
                        "productName": "单门 0.8锌合金 无造型 无造型",
                        "width": 1000,
                        "height": 2200,
                        "openDirection": "左内开",
                        "unit": "m2",
                        "unitPrice": 1200,
                    }
                ],
            },
        ],
    }


class MultiDoorQuoteTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory(prefix="door_multi_quote_")
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def quote_manager(self):
        return QuoteDatabaseManager(
            file_path=str(self.root / "quotes.json"),
            backup_dir=str(self.root / "backups"),
        )

    def test_multi_door_quote_persists_groups_and_flat_items(self):
        manager = self.quote_manager()
        created = manager.create(_multi_door_quote())
        loaded = manager.get_by_id(created["id"])

        self.assertEqual(len(loaded["doorGroups"]), 2)
        self.assertEqual(loaded["doorGroups"][0]["groupName"], "入户门")
        self.assertEqual(loaded["doorGroups"][1]["pricingMode"], "framePlusTrim")
        self.assertEqual(len(loaded["items"]), 3)
        self.assertEqual([item["groupIndex"] for item in loaded["items"]], [0, 0, 1])
        self.assertEqual(loaded["doorGroups"][0]["items"][1]["openDirection"], "")

    def test_legacy_quote_is_wrapped_as_one_door_group(self):
        manager = self.quote_manager()
        created = manager.create({
            "customerName": "旧报价客户",
            "projectName": "",
            "quoteDate": "2026-07-23",
            "items": [{"productName": "旧报价门", "unit": "套", "unitPrice": 1000}],
        })

        stored = manager._load_unlocked()
        stored[0].pop("doorGroups", None)
        manager._atomic_save(stored)
        loaded = manager.get_by_id(created["id"])

        self.assertEqual(len(loaded["doorGroups"]), 1)
        self.assertEqual(loaded["doorGroups"][0]["items"][0]["productName"], "旧报价门")

    def test_multi_door_excel_uses_group_and_total_formulas(self):
        quote = self.quote_manager().create(_multi_door_quote())
        output_path = self.root / "multi-door.xlsx"
        generate_excel(quote, str(output_path))
        workbook = load_workbook(output_path, data_only=False)
        sheet = workbook["Sheet1 (2)"] if "Sheet1 (2)" in workbook.sheetnames else workbook.worksheets[0]

        values = [
            sheet.cell(row, column).value
            for row in range(9, 17)
            for column in range(1, 11)
        ]
        self.assertNotIn("入户门", values)
        self.assertNotIn("厨房门", values)
        self.assertEqual(sheet["A11"].value, "入户门小计")
        self.assertEqual(sheet["J11"].value, "=SUM(J9,J10)")
        self.assertEqual(sheet["A13"].value, "厨房门小计")
        self.assertEqual(sheet["J13"].value, "=SUM(J12)")
        self.assertEqual(sheet["J17"].value, "=SUM(J11,J13)")
        self.assertIn("J17", sheet["F18"].value)
        self.assertIn("A1:J24", str(sheet.print_area).replace("$", ""))

    def test_multi_door_html_hides_group_titles_but_keeps_subtotals(self):
        quote_path = self.root / "quote.json"
        quote_path.write_text(json.dumps(_multi_door_quote(), ensure_ascii=False), encoding="utf-8")
        repo_root = Path(BACKEND_DIR).parent
        script = """
import { buildQuoteHtml } from './quote-template-pdf/src/renderQuote.mjs';
const html = await buildQuoteHtml(process.argv[1]);
process.stdout.write(html);
"""
        result = subprocess.run(
            ["node", "--input-type=module", "-e", script, str(quote_path)],
            cwd=repo_root,
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=False,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertNotIn('<tr class="group-row">', result.stdout)
        self.assertIn("入户门小计", result.stdout)
        self.assertIn("厨房门小计", result.stdout)

    def test_quote_memory_upserts_name_category_unit_and_price(self):
        manager = AccessoryDatabaseManager(
            file_path=str(self.root / "accessories.json"),
            backup_dir=str(self.root / "backups"),
        )
        memory_row = {
            "name": "暗合页（6只）",
            "category": "合页",
            "unit": "套",
            "unitPrice": 1000,
        }

        self.assertEqual(manager.import_batch([memory_row]), 1)
        self.assertEqual(manager.import_batch([{**memory_row, "unitPrice": 1100}]), 1)
        matched = [
            item for item in manager.get_all()
            if item["name"] == memory_row["name"] and item["category"] == memory_row["category"]
        ]

        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["unit"], "套")
        self.assertEqual(matched[0]["unitPrice"], 1100)

    def test_cad_cache_reuses_generated_dxf(self):
        import main

        calls = {"count": 0}

        def fake_build(_request):
            return {}, {}, {}

        def fake_generate(_info, _checks, _params, _progress):
            calls["count"] += 1
            return "ok", io.StringIO("SECTION\nENDSEC\nEOF")

        main._cad_cache.clear()
        request = CADRequest(ddh="CACHE-TEST")
        with patch.object(main, "build_cad_params", fake_build), patch.object(
            main, "run_integrated_system", fake_generate
        ):
            first_key, first_bytes, first_hit = main._cached_cad(request)
            second_key, second_bytes, second_hit = main._cached_cad(request)

        self.assertEqual(calls["count"], 1)
        self.assertFalse(first_hit)
        self.assertTrue(second_hit)
        self.assertEqual(second_key, first_key)
        self.assertEqual(second_bytes, first_bytes)


if __name__ == "__main__":
    unittest.main()
