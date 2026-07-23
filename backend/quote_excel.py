"""
报价单 Excel 生成 + JPG 渲染
使用 openpyxl 将报价数据写入 template.xlsx 模板
使用 LibreOffice + poppler-utils 将 Excel A1:J24 精确渲染为 JPG
"""
import os
import math
import subprocess
import shutil
import tempfile
import unicodedata
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image


def _resolve_project_root(module_file: str = __file__) -> Path:
    module_dir = Path(module_file).resolve().parent
    if module_dir.name == "backend":
        return module_dir.parent
    return module_dir


TEMPLATE_PATH = str(_resolve_project_root() / "template.xlsx")
DEFAULT_NOTICE_TEXT = "\u672c\u62a5\u4ef7\u4e0d\u542b\u7a0e\u5de5\u5382\u7ed3\u7b97\u4ef7\uff0c\u542b\u6728\u7bb1\u3002"


def _display_width(value) -> int:
    text = "" if value is None else str(value)
    return sum(2 if unicodedata.east_asian_width(char) in ("W", "F", "A") else 1 for char in text)


def _column_width(ws, column: str) -> float:
    return float(ws.column_dimensions[column].width or 13)


def _line_count(value, capacity: float) -> int:
    text = "" if value is None else str(value)
    if not text:
        return 1
    usable = max(1, int(capacity))
    return max(1, sum(max(1, math.ceil(_display_width(line) / usable)) for line in text.splitlines() or [""]))


def _merged_text_capacity(ws, columns: str, utilization: float = 0.72) -> float:
    """Estimate usable text width inside merged cells with Chinese content."""
    raw_width = sum(_column_width(ws, column) for column in columns)
    return max(1, raw_width * utilization)


def _set_song_font(cell) -> None:
    font = copy(cell.font)
    font.name = "宋体"
    font.charset = 134
    cell.font = font


def _enable_wrap(cell) -> None:
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "center"
    cell.alignment = alignment


def _fit_row(ws, row: int, value, capacity: float, minimum: float, line_height: float = 18) -> None:
    ws.row_dimensions[row].height = max(minimum, _line_count(value, capacity) * line_height + 5)


def _fit_row_group(ws, rows: range, value, capacity: float, minimum_total: float, line_height: float = 18) -> None:
    total = max(minimum_total, _line_count(value, capacity) * line_height + 5)
    each = total / len(rows)
    for row in rows:
        ws.row_dimensions[row].height = each


def _apply_dynamic_layout(
    ws,
    quote: dict,
    items: list[dict],
    item_rows: list[int],
    total_row: int,
) -> None:
    product_width = max((_display_width(item.get("productName") or item.get("product_name") or "") for item in items), default=0)
    current_product_capacity = _column_width(ws, "B") + _column_width(ws, "C")
    if product_width > current_product_capacity:
        target_capacity = min(current_product_capacity + 6, max(current_product_capacity, product_width))
        ws.column_dimensions["C"].width = _column_width(ws, "C") + (target_capacity - current_product_capacity)

    product_capacity = _merged_text_capacity(ws, "BC")
    for row in item_rows:
        cell = ws[f"B{row}"]
        _enable_wrap(cell)
        _fit_row(ws, row, cell.value, product_capacity, 25, line_height=22)

    ws.column_dimensions["H"].width = max(_column_width(ws, "H"), 11.5)
    ws.column_dimensions["I"].width = max(_column_width(ws, "I"), 9.5)
    for row in item_rows:
        ws[f"H{row}"].number_format = "0.####"
        ws[f"J{row}"].number_format = "0"
    ws[f"J{total_row}"].number_format = "0"

    customer_capacity = sum(_column_width(ws, column) for column in ("C", "D", "E", "F"))
    for row, value in ((3, quote.get("customerName", "")), (4, quote.get("projectName", ""))):
        _enable_wrap(ws[f"C{row}"])
        _fit_row(ws, row, value, customer_capacity, 20.25, line_height=18)

    full_capacity = sum(_column_width(ws, column) for column in "ABCDEFGHIJ")
    notice_row = total_row + 2
    terms_row = total_row + 3
    invoice_row = total_row + 6
    bank_row = total_row + 7
    _enable_wrap(ws[f"A{notice_row}"])
    _fit_row(ws, notice_row, ws[f"A{notice_row}"].value, full_capacity, 34, line_height=18)
    _enable_wrap(ws[f"A{terms_row}"])
    _fit_row_group(ws, range(terms_row, terms_row + 3), ws[f"A{terms_row}"].value, full_capacity, 66, line_height=19)
    for row, minimum in ((invoice_row, 105), (bank_row, 81)):
        _enable_wrap(ws[f"A{row}"])
        _fit_row(ws, row, ws[f"A{row}"].value, full_capacity, minimum, line_height=19)

    amount_row = total_row + 1
    _enable_wrap(ws[f"F{amount_row}"])
    amount_capacity = sum(_column_width(ws, column) for column in "FGHIJ")
    _fit_row(ws, amount_row, ws[f"F{amount_row}"].value, amount_capacity, 20.25, line_height=18)

    for row in ws.iter_rows(min_row=1, max_row=bank_row, min_col=1, max_col=10):
        for cell in row:
            _set_song_font(cell)


def _quote_groups(quote: dict) -> list[dict]:
    groups = quote.get("doorGroups") or []
    if groups:
        return groups
    return [{
        "groupName": "第1樘门",
        "taskId": "",
        "pricingMode": "outerArea",
        "trimUnitPrice": 0,
        "items": quote.get("items") or [],
    }]


def _copy_row_format(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, 11):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)


def _unmerge_dynamic_rows(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 9:
            ws.unmerge_cells(str(merged_range))


def _is_area_unit(unit: str) -> bool:
    normalized = (unit or "").lower()
    return "m2" in normalized or "㎡" in normalized or "m²" in normalized


def _quote_quantity(item: dict) -> float:
    if not (item.get("productName") or item.get("product_name") or "").strip():
        return 0
    explicit_qty = item.get("quantity")
    if explicit_qty is not None and explicit_qty != "":
        try:
            return float(explicit_qty)
        except (TypeError, ValueError):
            pass
    if not _is_area_unit(item.get("unit") or ""):
        return 1
    width = float(item.get("width") or 0)
    height = float(item.get("height") or 0)
    if not width or not height:
        return 0
    return width * height * 0.000001


def _quote_amount(item: dict) -> int:
    qty = _quote_quantity(item)
    unit_price = float(item.get("unitPrice") or item.get("unit_price") or 0)
    if not qty:
        return 0
    return round(qty * unit_price)


def _to_chinese_amount(value: float) -> str:
    n = int(round(value or 0))
    if n <= 0:
        return ""
    digits = ["\u96f6", "\u58f9", "\u8d30", "\u53c1", "\u8086", "\u4f0d", "\u9646", "\u67d2", "\u634c", "\u7396"]
    units = ["", "\u62fe", "\u4f70", "\u4edf"]
    sections = ["", "\u4e07", "\u4ebf", "\u4e07\u4ebf"]

    def section_to_chinese(section: int) -> str:
        text = ""
        zero = False
        for i in range(4):
            divisor = 10 ** (3 - i)
            digit = (section // divisor) % 10
            unit_index = 3 - i
            if digit == 0:
                zero = bool(text)
            else:
                if zero:
                    text += digits[0]
                text += digits[digit] + units[unit_index]
                zero = False
        return text

    remaining = n
    section_index = 0
    result = ""
    need_zero = False
    while remaining > 0:
        section = remaining % 10000
        if section == 0:
            need_zero = bool(result)
        else:
            section_text = section_to_chinese(section) + sections[section_index]
            if need_zero or (section < 1000 and remaining >= 10000):
                section_text = digits[0] + section_text
            result = section_text + result
            need_zero = False
        remaining //= 10000
        section_index += 1
    return f"\u4eba\u6c11\u5e01{result}\u5143\u6574"


def generate_excel(quote: dict, output_path: str):
    """
    将报价数据写入 Excel 模板并保存到 output_path。
    quote: QuoteResponse dict（包含 customerName, projectName, quoteDate, items）
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1 (2)"] if "Sheet1 (2)" in wb.sheetnames else wb.worksheets[0]

    # Header
    ws["C3"] = quote.get("customerName", "")
    ws["C4"] = quote.get("projectName", "")
    ws["I3"] = quote.get("quoteDate", "")

    groups = _quote_groups(quote)
    display_rows: list[tuple[str, dict, int]] = []
    items: list[dict] = []
    for group_index, group in enumerate(groups):
        group_items = list(group.get("items") or [])
        if len(groups) > 1:
            display_rows.append(("group", group, group_index))
        for item in group_items:
            display_rows.append(("item", item, group_index))
            items.append(item)
        if len(groups) > 1:
            display_rows.append(("subtotal", group, group_index))
    while len(display_rows) < 8:
        display_rows.append(("item", {}, -1))

    extra_rows = max(0, len(display_rows) - 8)
    _unmerge_dynamic_rows(ws)
    if extra_rows:
        ws.insert_rows(17, amount=extra_rows)

    total_row = 9 + len(display_rows)
    amount_row = total_row + 1
    notice_row = total_row + 2
    terms_row = total_row + 3
    invoice_row = total_row + 6
    bank_row = total_row + 7
    item_rows: list[int] = []
    group_item_rows: dict[int, list[int]] = {}
    sequence = 0

    for row_offset, (row_type, payload, marker) in enumerate(display_rows):
        row = 9 + row_offset
        if row > 16:
            _copy_row_format(ws, 16, row)
        for column in range(1, 11):
            ws.cell(row, column).value = None

        if row_type == "group":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            ws[f"A{row}"] = payload.get("groupName") or f"第{marker + 1}樘门"
            ws[f"A{row}"].fill = PatternFill("solid", fgColor="F2F2F7")
            font = copy(ws[f"A{row}"].font)
            font.bold = True
            ws[f"A{row}"].font = font
            continue

        if row_type == "subtotal":
            _copy_row_format(ws, total_row, row)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            group_rows = group_item_rows.get(marker, [])
            ws[f"A{row}"] = f'{payload.get("groupName") or f"第{marker + 1}樘门"}小计'
            ws[f"J{row}"] = f"=SUM({','.join(f'J{item_row}' for item_row in group_rows)})" if group_rows else "=0"
            continue

        item = payload
        item_rows.append(row)
        group_index = marker if marker >= 0 else 0
        group_item_rows.setdefault(group_index, []).append(row)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        product_name = item.get("productName") or item.get("product_name", "")
        if product_name:
            sequence += 1
            ws[f"A{row}"] = sequence
        ws[f"B{row}"] = product_name
        ws[f"D{row}"] = item.get("width")
        ws[f"E{row}"] = item.get("height")
        ws[f"F{row}"] = item.get("openDirection") or item.get("open_direction", "")
        ws[f"G{row}"] = item.get("unit") or ""
        if item.get("quantity") is not None:
            ws[f"H{row}"] = item.get("quantity")
        else:
            ws[f"H{row}"] = f'=IF(B{row}="","",IF(OR(G{row}="m2",G{row}="㎡",G{row}="m²"),IF(OR(D{row}="",E{row}=""),"",D{row}*E{row}*0.000001),1))'
        ws[f"I{row}"] = item.get("unitPrice") or item.get("unit_price", 0)
        ws[f"J{row}"] = f'=IF(OR(H{row}="",I{row}=""),"",ROUND(H{row}*I{row},0))'

    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    ws[f"A{total_row}"] = "合计"
    subtotal_rows = [9 + index for index, entry in enumerate(display_rows) if entry[0] == "subtotal"]
    if subtotal_rows:
        ws[f"J{total_row}"] = f"=SUM({','.join(f'J{row}' for row in subtotal_rows)})"
    else:
        ws[f"J{total_row}"] = f"=SUM(J{min(item_rows)}:J{max(item_rows)})"

    ws.merge_cells(start_row=amount_row, start_column=1, end_row=amount_row, end_column=5)
    ws.merge_cells(start_row=amount_row, start_column=6, end_row=amount_row, end_column=10)
    ws[f"A{amount_row}"] = "合计总金额（大写）："
    ws[f"F{amount_row}"] = f'=IF(J{total_row}=0,"","人民币"&TEXT(ROUND(J{total_row},0),"[DBNum2]")&"元整")'
    ws.merge_cells(start_row=notice_row, start_column=1, end_row=notice_row, end_column=10)
    ws[f"A{notice_row}"] = quote.get("noticeText") or DEFAULT_NOTICE_TEXT
    ws.merge_cells(start_row=terms_row, start_column=1, end_row=terms_row + 2, end_column=10)
    ws.merge_cells(start_row=invoice_row, start_column=1, end_row=invoice_row, end_column=10)
    ws.merge_cells(start_row=bank_row, start_column=1, end_row=bank_row, end_column=10)

    _apply_dynamic_layout(ws, quote, items, item_rows, total_row)

    ws.print_area = f"A1:J{bank_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output_path)


# ===================== JPG 渲染：Excel A1:J24 → 图片 =====================

def render_quote_jpg(quote: dict, output_path: str):
    """
    将报价数据渲染为 JPG 图片（Excel A1:J24 + 40px 白边）。
    管线：openpyxl 生成 Excel → LibreOffice 渲染为 PDF → pdftoppm 转 JPG → Pillow 加白边。
    输出与 Excel 原生导出 A1:J24 完全一致。
    """
    work_dir = tempfile.mkdtemp()

    try:
        # 1. 生成 Excel
        xlsx_path = os.path.join(work_dir, "quote.xlsx")
        generate_excel(quote, xlsx_path)

        # 2. LibreOffice headless: xlsx → PDF
        subprocess.run(
            ["libreoffice", "--headless", "--norestore", "--convert-to", "pdf",
             "--outdir", work_dir, xlsx_path],
            check=True, timeout=60, capture_output=True,
        )

        # 找到输出的 PDF 文件
        pdf_files = [f for f in os.listdir(work_dir) if f.endswith(".pdf")]
        if not pdf_files:
            raise RuntimeError("LibreOffice 未生成 PDF")
        pdf_path = os.path.join(work_dir, pdf_files[0])

        # 3. pdftoppm: PDF → JPEG（200 DPI）
        jpg_base = os.path.join(work_dir, "page")
        subprocess.run(
            ["pdftoppm", "-jpeg", "-r", "200", "-singlefile", pdf_path, jpg_base],
            check=True, timeout=30, capture_output=True,
        )
        raw_jpg = jpg_base + ".jpg"
        if not os.path.exists(raw_jpg):
            raise RuntimeError("pdftoppm 未生成 JPG")

        # 4. Pillow: 加 40px 白边
        with Image.open(raw_jpg) as img:
            margin = 40
            new_w = img.width + 2 * margin
            new_h = img.height + 2 * margin
            out = Image.new("RGB", (new_w, new_h), "#FFFFFF")
            out.paste(img, (margin, margin))
            out.save(output_path, "JPEG", quality=92)

    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


def render_quote_pdf(quote: dict, output_path: str):
    """
    灏嗘姤浠锋暟鎹覆鏌撲负 PDF 鏂囦欢锛屽鐢?JPG 娓叉煋绠＄嚎锛屼究浜庝簯绔ǔ瀹氬鍑恒€?
    """
    work_dir = tempfile.mkdtemp()

    try:
        jpg_path = os.path.join(work_dir, "quote.jpg")
        render_quote_jpg(quote, jpg_path)

        with Image.open(jpg_path) as img:
            img.convert("RGB").save(output_path, "PDF", resolution=200.0)
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)
