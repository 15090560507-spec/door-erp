"""
报价系统 API 路由
"""
import logging
import os
import shutil
import tempfile
import base64
import json
import re
import time
import urllib.error
import urllib.request
from io import BytesIO
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from starlette.background import BackgroundTask
from openpyxl import load_workbook
from config import DATA_DIR
from quote_models import (
    AccessoryCreate, AccessoryImport,
    QuoteCreate,
    AiConfigUpdate,
)
from quote_database import AccessoryDatabaseManager, QuoteDatabaseManager, AiConfigManager
from quote_excel import generate_excel
from quote_template_renderer import render_quote_template_artifacts

quote_router = APIRouter()
logger = logging.getLogger(__name__)
AI_MODEL_ALIASES = {
    "kimi-k2.5": "moonshot-v1-8k-vision-preview",
    "kimi-k2.6": "moonshot-v1-8k-vision-preview",
    "kimi-k2.65": "moonshot-v1-8k-vision-preview",
}

# 实例化管理器
accessory_db = AccessoryDatabaseManager()
quote_db = QuoteDatabaseManager()
ai_config_db = AiConfigManager()


def _cleanup_dir(path: str):
    shutil.rmtree(path, ignore_errors=True)


def _extract_json_payload(text: str) -> dict:
    cleaned = (text or "").strip()
    if not cleaned:
        raise HTTPException(
            status_code=502,
            detail="AI 返回内容为空；请检查模型是否支持图片识别，建议使用 moonshot-v1-8k-vision-preview。",
        )
    fence_match = re.search(r"```(?:json)?\s*(.*?)```", cleaned, re.S | re.I)
    if fence_match:
        cleaned = fence_match.group(1).strip()
    if not cleaned.startswith("{"):
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            cleaned = cleaned[start:end + 1]
    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        preview = cleaned[:200].replace("\n", " ")
        raise HTTPException(status_code=502, detail=f"AI 返回内容不是有效 JSON: {exc}；返回开头：{preview}") from exc
    if not isinstance(data, dict):
        raise HTTPException(status_code=502, detail="AI 返回 JSON 顶层必须是对象")
    return data


def _number_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        match = re.search(r"-?\d+(?:\.\d+)?", value)
        if match:
            return float(match.group(0))
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_analysis(data: dict) -> dict:
    items = data.get("items") or []
    if not isinstance(items, list):
        items = []
    normalized_items = []
    for item in items[:8]:
        if not isinstance(item, dict):
            continue
        normalized_items.append({
            "productName": str(item.get("productName") or item.get("name") or ""),
            "width": _number_or_none(item.get("width")),
            "height": _number_or_none(item.get("height")),
            "quantity": _number_or_none(item.get("quantity")),
            "openDirection": str(item.get("openDirection") or ""),
            "unit": str(item.get("unit") or "m2"),
            "unitPrice": _number_or_none(item.get("unitPrice")),
        })

    accessories = data.get("accessories") or []
    if isinstance(accessories, str):
        accessories = [accessories]
    if not isinstance(accessories, list):
        accessories = []

    return {
        "customerName": str(data.get("customerName") or ""),
        "projectName": str(data.get("projectName") or ""),
        "outerWidth": _number_or_none(data.get("outerWidth")),
        "outerHeight": _number_or_none(data.get("outerHeight")),
        "openDirection": str(data.get("openDirection") or ""),
        "items": normalized_items,
        "accessories": [str(item) for item in accessories if item],
        "notes": str(data.get("notes") or ""),
    }


def _normalize_ai_model(model: str) -> str:
    normalized = (model or "").strip()
    return AI_MODEL_ALIASES.get(normalized, normalized)


def _prepare_ai_image(image_bytes: bytes, content_type: str) -> tuple[bytes, str]:
    """Keep uploaded drawings unmodified so OCR sees the original details."""
    return image_bytes, content_type


def _parse_price(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else 0.0


def _clean_cell(value) -> str:
    return str(value or "").strip()


def _strip_note(value: str) -> str:
    return re.sub(r"[（(].*?[）)]", "", value).strip()


def _append_option(options: dict, key: str, value: str):
    value = _clean_cell(value)
    if not value:
        return
    existing = options.setdefault(key, [])
    if value not in existing:
        existing.append(value)


def _sync_dropdown_options_from_price_items(items: list[dict]):
    path = os.path.join(DATA_DIR, "dropdown_options.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            options = json.load(f)
    except Exception:
        options = {}

    material_items = [item for item in items if item.get("category", "") == "制作材料"]
    if material_items:
        options["MATERIALS"] = []

    for item in items:
        category = item.get("category", "")
        if category == "制作材料":
            _append_option(options, "MATERIALS", item.get("name", ""))
        elif category == "款式组合":
            _append_option(options, "DOOR_STYLES", item.get("frontStyle", ""))
            _append_option(options, "DOOR_STYLES", item.get("backStyle", ""))
        elif category == "锁体":
            _append_option(options, "LOCKS", item.get("name", ""))
        elif category == "合页":
            _append_option(options, "HINGES", _strip_note(item.get("name", "")))
        elif category == "包装":
            _append_option(options, "BZ_OPTIONS", item.get("name", ""))
        elif category == "配件" and "指纹锁" in item.get("name", ""):
            _append_option(options, "FINGERPRINT_LOCKS", item.get("name", ""))

    fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".json")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(options, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


def _parse_accessory_price_xlsx(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 文件读取失败: {exc}") from exc

    sheet = workbook.worksheets[0]
    items: list[dict] = []
    seen: set[tuple[str, str, str, str]] = set()

    def add_item(item: dict):
        name = _clean_cell(item.get("name"))
        if not name:
            return
        item["name"] = name
        key = (
            _clean_cell(item.get("category")),
            name,
            _clean_cell(item.get("frontStyle")),
            _clean_cell(item.get("backStyle")),
        )
        if key in seen:
            return
        seen.add(key)
        items.append(item)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row) + [None] * 13

        material = _clean_cell(values[0])
        if material:
            add_item({
                "name": material,
                "category": "制作材料",
                "keywords": material,
                "unit": "m2",
                "unitPrice": _parse_price(values[1]),
                "remark": "面积单价加价",
                "priceType": "material",
                "priceMode": "area_add",
            })

        front_style = _clean_cell(values[2])
        back_style = _clean_cell(values[3])
        if front_style and back_style and values[4] not in (None, "") and front_style != "正面" and back_style != "反面":
            add_item({
                "name": f"{front_style}+{back_style}",
                "category": "款式组合",
                "keywords": f"{front_style} {back_style}",
                "unit": "m2",
                "unitPrice": _parse_price(values[4]),
                "remark": "面积基础单价",
                "priceType": "style_combo",
                "priceMode": "area_base",
                "frontStyle": front_style,
                "backStyle": back_style,
            })

        lock_body = _clean_cell(values[5])
        if lock_body:
            add_item({
                "name": lock_body,
                "category": "锁体",
                "keywords": lock_body,
                "unit": "m2",
                "unitPrice": _parse_price(values[6]),
                "remark": "面积单价加价",
                "priceType": "lock_body",
                "priceMode": "area_add",
            })

        hinge = _clean_cell(values[7])
        if hinge:
            add_item({
                "name": hinge,
                "category": "合页",
                "keywords": f"{hinge} {_strip_note(hinge)}",
                "unit": "套",
                "unitPrice": _parse_price(values[8]),
                "remark": "单独计价",
                "priceType": "hardware",
                "priceMode": "piece",
            })

        accessory = _clean_cell(values[9])
        if accessory:
            price = 580.0 if accessory.upper().startswith("Q3") else _parse_price(values[10])
            add_item({
                "name": accessory,
                "category": "配件",
                "keywords": accessory,
                "unit": "套",
                "unitPrice": price,
                "remark": "单独计价",
                "priceType": "hardware",
                "priceMode": "piece",
            })

        packing = _clean_cell(values[11])
        if packing:
            add_item({
                "name": packing,
                "category": "包装",
                "keywords": packing,
                "unit": "m2",
                "unitPrice": _parse_price(values[12]),
                "remark": "按外围面积计价",
                "priceType": "packing",
                "priceMode": "outer_area_add",
            })

    return items


# ===================== 配件管理 =====================

@quote_router.get("/api/accessories")
def list_accessories(q: str = Query(None, description="搜索关键词")):
    """获取配件列表，支持按名称/类别/型号/关键词搜索"""
    if q:
        accessories = accessory_db.search(q)
    else:
        accessories = accessory_db.get_all()
    return {"accessories": accessories}


@quote_router.post("/api/accessories", status_code=201)
def create_accessory(data: AccessoryCreate):
    """新增配件"""
    if not data.name.strip():
        raise HTTPException(status_code=400, detail="配件名称不能为空")
    try:
        new_acc = accessory_db.add(data.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"id": new_acc["id"]}


# NOTE: export/import 路由必须放在 /{accessory_id} 之前，避免 FastAPI 路由冲突
@quote_router.get("/api/accessories/export")
def export_accessories():
    """导出所有有效配件为 JSON"""
    accessories = accessory_db.get_all()
    return {"accessories": accessories}


@quote_router.post("/api/accessories/import", status_code=201)
def import_accessories(data: AccessoryImport):
    """批量导入配件"""
    items = [item.model_dump() for item in data.accessories]
    count = accessory_db.import_batch(items)
    return {"imported": count}


@quote_router.post("/api/accessories/import-xlsx", status_code=201)
async def import_accessories_xlsx(file: UploadFile = File(...)):
    """Import the door price workbook into the accessory/price library."""
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件")
    content = await file.read()
    items = _parse_accessory_price_xlsx(content)
    if not items:
        raise HTTPException(status_code=400, detail="Excel 中没有识别到可导入的价格数据")
    count = accessory_db.import_batch(items)
    _sync_dropdown_options_from_price_items(items)
    return {"imported": count, "parsed": len(items)}


@quote_router.delete("/api/accessories/{accessory_id}")
def delete_accessory(accessory_id: int):
    """软删除配件（将 active 置为 0）"""
    try:
        accessory_db.soft_delete(accessory_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"deleted": accessory_id}


# ===================== 报价单管理 =====================

@quote_router.get("/api/quotes")
def list_quotes():
    """获取报价单列表（最新 50 条，不含 items 明细以优化性能）"""
    quotes = quote_db.get_all(limit=50)
    return {"quotes": quotes}


@quote_router.post("/api/quotes", status_code=201)
def create_quote(data: QuoteCreate):
    """创建报价单"""
    try:
        quote = quote_db.create(data.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"quote": quote}


@quote_router.get("/api/quotes/{quote_id}")
def get_quote(quote_id: int):
    """获取单个报价单详情（含 items）"""
    quote = quote_db.get_by_id(quote_id)
    if not quote:
        raise HTTPException(status_code=404, detail="报价单不存在")
    return {"quote": quote}


@quote_router.delete("/api/quotes/{quote_id}")
def delete_quote(quote_id: int):
    """删除报价单"""
    try:
        quote_db.delete(quote_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail="报价单不存在")
    return {"deleted": quote_id}


# ===================== 报价单导出 =====================

@quote_router.get("/api/quotes/{quote_id}/export.xlsx")
def export_quote_xlsx(quote_id: int):
    """导出报价单为 Excel"""
    quote = quote_db.get_by_id(quote_id)
    if not quote:
        raise HTTPException(status_code=404, detail="报价单不存在")

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        generate_excel(quote, tmp_path)
        filename = f"quote_{quote_id}.xlsx"
        return FileResponse(
            tmp_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            background=BackgroundTask(lambda path: os.path.exists(path) and os.unlink(path), tmp_path),
        )
    except Exception:
        logger.exception("Excel export failed for quote_id=%s", quote_id)
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise HTTPException(status_code=500, detail="Excel 生成失败")


def _build_quote_html(quote: dict, auto_print: bool = False) -> str:
    """构建报价单打印 HTML，仅使用标准 hex/rgb 颜色（兼容 html2canvas）"""
    items_html = ""
    items = quote.get("items", [])[:8]
    for i, item in enumerate(items):
        width = item.get("width") or ""
        height = item.get("height") or ""
        unit_price = item.get("unitPrice") or ""
        explicit_qty = item.get("quantity")
        qty = ""
        amount = ""
        unit = item.get("unit") or ""
        is_area_unit = "m2" in str(unit).lower() or "㎡" in str(unit) or "m²" in str(unit)
        if explicit_qty not in (None, ""):
            q = float(explicit_qty)
            qty = f"{q:.4f}" if is_area_unit else str(q)
            if unit_price:
                amount = str(round(q * float(unit_price)))
        elif item.get("productName") and not is_area_unit:
            qty = "1"
            if unit_price:
                amount = str(round(float(unit_price)))
        elif width and height:
            q = float(width) * float(height) * 0.000001
            qty = f"{q:.4f}"
            if unit_price:
                amount = str(round(q * float(unit_price)))
        items_html += f"""<tr>
            <td>{i + 1}</td>
            <td>{item.get('productName', '')}</td>
            <td>{width}</td><td>{height}</td>
            <td>{item.get('openDirection', '') if i == 0 else ''}</td>
            <td>{item.get('unit', '')}</td>
            <td>{qty}</td><td>{unit_price}</td><td>{amount}</td>
        </tr>"""

    auto_print_script = "<script>window.onload=function(){window.print()}</script>" if auto_print else ""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="utf-8"><title>报价单</title>
<style>
  @page {{ size: A4; margin: 12mm; }}
  body {{ font-family: "PingFang SC", "Microsoft YaHei", sans-serif; font-size: 13px; color: #1C1C1E; background: #FFFFFF; margin: 20px; }}
  h2 {{ text-align: center; font-size: 18px; margin-bottom: 12px; color: #1C1C1E; }}
  .meta {{ display: grid; grid-template-columns: 1fr 1fr; gap: 4px 24px; margin-bottom: 10px; }}
  .meta div {{ font-size: 13px; color: #1C1C1E; }}
  .meta strong {{ color: #8E8E93; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 10px; font-size: 11px; }}
  th, td {{ border: 1px solid #333; padding: 4px 3px; text-align: center; color: #1C1C1E; }}
  th {{ background: #F5F5F5; font-weight: bold; }}
  .terms {{ font-size: 11px; color: #555555; margin-top: 10px; }}
  .terms p {{ color: #555555; margin: 2px 0; }}
</style></head>
<body>
<h2>浙江西州将军门业有限公司</h2>
<div class="meta">
  <div><strong>客户名称：</strong>{quote.get('customerName', '')}</div>
  <div><strong>日期：</strong>{quote.get('quoteDate', '')}</div>
  <div><strong>项目名称：</strong>{quote.get('projectName', '')}</div>
  <div><strong>主题：</strong>产品报价单</div>
</div>
<table>
<thead>
<tr><th rowspan="2">序号</th><th rowspan="2">品名型号</th><th colspan="2">规格</th><th rowspan="2">开启方向</th><th rowspan="2">单位</th><th rowspan="2">数量</th><th rowspan="2">单价</th><th rowspan="2">总金额/元</th></tr>
<tr><th>宽</th><th>高</th></tr>
</thead>
<tbody>{items_html}</tbody>
</table>
<div class="terms">
  <p>本报价不含税工厂结算价，不含木箱。</p>
  <p>1. 付款方式：确定制作，先安排货款 50% 的定金，款清发货。</p>
  <p>2. 费用说明：以上价格不包含运输、安装、测量等费用。</p>
  <p>3. 确认流程：请及时确认签字回传，以便安排生产。</p>
</div>
{auto_print_script}
</body></html>"""


@quote_router.get("/api/quotes/{quote_id}/preview.html")
def quote_preview_html(quote_id: int):
    """返回报价单预览 HTML，由 Playwright 模板渲染。"""
    quote = quote_db.get_by_id(quote_id)
    if not quote:
        raise HTTPException(status_code=404, detail="报价单不存在")

    tmp_dir = tempfile.mkdtemp(prefix="quote-preview-")
    try:
        artifacts = render_quote_template_artifacts(quote, tmp_dir)
        with open(artifacts["html_path"], "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    except Exception:
        logger.exception("Quote preview render failed for quote_id=%s", quote_id)
        _cleanup_dir(tmp_dir)
        raise HTTPException(status_code=500, detail="HTML 预览生成失败")
    finally:
        _cleanup_dir(tmp_dir)


@quote_router.get("/api/quotes/{quote_id}/export.jpg")
def export_quote_jpg(quote_id: int):
    """导出报价单 JPG，使用高还原 HTML/CSS 模板渲染。"""
    quote = quote_db.get_by_id(quote_id)
    if not quote:
        raise HTTPException(status_code=404, detail="报价单不存在")

    tmp_dir = tempfile.mkdtemp(prefix="quote-jpg-")
    try:
        artifacts = render_quote_template_artifacts(quote, tmp_dir)
        return FileResponse(
            artifacts["jpg_path"],
            media_type="image/jpeg",
            filename=f"quote_{quote_id}.jpg",
            background=BackgroundTask(_cleanup_dir, tmp_dir),
        )
    except Exception:
        logger.exception("JPG export failed for quote_id=%s", quote_id)
        _cleanup_dir(tmp_dir)
        raise HTTPException(status_code=500, detail="JPG 生成失败")


@quote_router.get("/api/quotes/{quote_id}/export.pdf")
def export_quote_pdf(quote_id: int):
    """导出报价单 PDF，使用高还原 HTML/CSS 模板渲染。"""
    quote = quote_db.get_by_id(quote_id)
    if not quote:
        raise HTTPException(status_code=404, detail="报价单不存在")

    tmp_dir = tempfile.mkdtemp(prefix="quote-pdf-")
    try:
        artifacts = render_quote_template_artifacts(quote, tmp_dir)
        return FileResponse(
            artifacts["pdf_path"],
            media_type="application/pdf",
            filename=f"quote_{quote_id}.pdf",
            background=BackgroundTask(_cleanup_dir, tmp_dir),
        )
    except Exception:
        logger.exception("PDF export failed for quote_id=%s", quote_id)
        _cleanup_dir(tmp_dir)
        raise HTTPException(status_code=500, detail="PDF 生成失败")


# ===================== AI 配置管理 =====================

@quote_router.get("/api/ai-config")
def get_ai_config():
    """获取 AI 识别配置"""
    config = ai_config_db.get()
    return {"config": config}


@quote_router.post("/api/drawings/analyze")
async def analyze_drawing(file: UploadFile = File(...)):
    """调用兼容 OpenAI chat/completions 的视觉模型识别报价图纸。"""
    content_type = file.content_type or ""
    if content_type not in {"image/jpeg", "image/png"}:
        raise HTTPException(status_code=400, detail="仅支持 JPG、JPEG、PNG 图纸")

    config = ai_config_db.get()
    base_url = (config.get("baseUrl") or "").strip().rstrip("/")
    endpoint_path = (config.get("endpointPath") or "/chat/completions").strip()
    api_key = (config.get("apiKey") or "").strip()
    model = _normalize_ai_model(config.get("model") or "")
    prompt = (config.get("prompt") or "").strip()
    if not base_url or not api_key or not model:
        raise HTTPException(status_code=400, detail="请先配置 AI Base URL、API Key 和模型名")

    image_bytes = await file.read()
    if not image_bytes:
        raise HTTPException(status_code=400, detail="上传图纸为空")

    image_bytes, content_type = _prepare_ai_image(image_bytes, content_type)
    data_url = f"data:{content_type};base64,{base64.b64encode(image_bytes).decode('ascii')}"
    url = f"{base_url}/{endpoint_path.lstrip('/')}"
    payload = {
        "model": model,
        "temperature": 1,
        "max_tokens": 1200,
        "messages": [
            {"role": "system", "content": "你是门业报价图纸识别助手。只输出 JSON，不要输出 Markdown 或解释。"},
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt or (
                            "请从图纸中提取报价需要的字段，输出 JSON："
                            "{\"customerName\":\"\",\"projectName\":\"\",\"outerWidth\":null,"
                            "\"outerHeight\":null,\"openDirection\":\"\",\"items\":[],"
                            "\"accessories\":[],\"notes\":\"\"}"
                        ),
                    },
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            },
        ],
    }
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        message = body[:300]
        try:
            error_payload = json.loads(body)
            error_info = error_payload.get("error", error_payload)
            if isinstance(error_info, dict):
                message = error_info.get("message") or error_info.get("detail") or message
        except json.JSONDecodeError:
            pass
        logger.warning("AI analysis HTTP error: %s %s", exc.code, body[:500])
        raise HTTPException(status_code=502, detail=f"AI 接口返回错误 {exc.code}: {message}") from exc
    except Exception as exc:
        logger.exception("AI analysis request failed")
        raise HTTPException(status_code=502, detail=f"AI 识别请求失败: {exc}") from exc

    try:
        content = response_data["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as exc:
        raise HTTPException(status_code=502, detail="AI 返回结构缺少 choices[0].message.content") from exc
    if isinstance(content, list):
        content = "\n".join(str(part.get("text", "")) if isinstance(part, dict) else str(part) for part in content)
    if not str(content).strip():
        raise HTTPException(
            status_code=502,
            detail=f"AI 返回内容为空；当前实际调用模型为 {model}，请确认该模型支持图片识别。",
        )

    analysis = _normalize_analysis(_extract_json_payload(str(content)))
    return {
        "uploadId": int(time.time() * 1000),
        "filename": file.filename or "drawing",
        "analysis": analysis,
        "rawPreview": str(content)[:1000],
    }


@quote_router.post("/api/ai-config")
def update_ai_config(data: AiConfigUpdate):
    """更新 AI 识别配置"""
    config = ai_config_db.update(data.model_dump())
    return {"config": config}


# ===================== 图纸分析（占位） =====================

@quote_router.post("/api/drawings/analyze-placeholder-disabled")
def analyze_drawing_placeholder():
    """AI 分析图纸（功能开发中）"""
    return {"message": "AI analysis coming soon"}
